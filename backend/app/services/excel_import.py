"""Excel-template import service (REQ-no-evidence-upload final form).

This is the only multipart-fed code path in the entire backend (Plan 06
locks the allow-list to a single endpoint:
``POST /api/v1/konkin/templates/{id}/import-from-excel``).

Pattern (RESEARCH.md Code Examples + Pitfall #10):

1. **Idempotency-Key gate.** If the caller sends the ``Idempotency-Key``
   header and ``konkin_import_log`` already contains a row with that key,
   short-circuit to ``status="already_applied"`` and DO NOT re-parse the
   workbook. Network retries / double-clicks become safe by construction.

2. **Streaming parse.** ``openpyxl.load_workbook(BytesIO(raw),
   read_only=True, data_only=True)``. ``read_only=True`` is the streaming
   mode (constant memory regardless of workbook size); ``data_only=True``
   short-circuits formula evaluation (we read computed values, not the
   formulas — the source-of-truth lives outside the sheet anyway).

3. **SAVEPOINT per sheet.** Each sheet is parsed inside its own
   ``async with db.begin_nested()`` so a partial parse failure
   (malformed row, integrity violation) rolls back JUST THAT SHEET, not
   the entire workbook. The outer transaction is committed once at the
   end so the import is atomic across sheets too.

4. **Phase-1 scope.** Sheet parsing returns 0 rows for now — Plan 06's
   Phase-1 scope is to BOOK the multipart contract, the idempotency log,
   the CSRF gate (B-07), and the size limits (Plan 02 Nginx
   ``client_max_body_size 25M``). Phase 2 will wire actual perspektif +
   indikator row creation; the test suite asserts the surface
   (status + already_applied semantics + log row) — not the row counts.

Public surface:

- ``ALLOWED_CT`` — content-type allow-list. Anything else returns 415 at
  the router layer.
- ``import_workbook(db, template, raw, idempotency_key, imported_by)`` —
  async function returning::

      {"status": "imported" | "already_applied",
       "log_id": "<uuid>" | None,
       "summary": {sheet_name: row_count, ...}}
"""

from __future__ import annotations

from io import BytesIO
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.konkin_import_log import KonkinImportLog
from app.models.konkin_template import KonkinTemplate

# Plan 06 final allow-list — keep in sync with
# `backend/tests/test_no_upload_policy.py::ALLOWED_MULTIPART_PATHS`.
ALLOWED_CT: set[str] = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
}


async def import_workbook(
    db: AsyncSession,
    template: KonkinTemplate,
    raw: bytes,
    idempotency_key: str | None,
    imported_by: UUID,
) -> dict:
    """Parse + ingest a workbook with idempotency + per-sheet SAVEPOINT.

    Returns a dict with one of two shapes::

        {"status": "already_applied", "log_id": "<uuid>", "summary": {...}}
        {"status": "imported",        "log_id": "<uuid>"|None, "summary": {...}}

    ``log_id`` is None when ``idempotency_key`` is not provided (no log row
    is written in that case — caller opted out of idempotency).
    """
    # ---------------------------------------------------------------------
    # 1. Idempotency gate — short-circuit on duplicate key.
    # ---------------------------------------------------------------------
    if idempotency_key:
        existing = await db.scalar(
            select(KonkinImportLog).where(
                KonkinImportLog.idempotency_key == idempotency_key
            )
        )
        if existing:
            return {
                "status": "already_applied",
                "log_id": str(existing.id),
                "summary": existing.summary,
            }

    # ---------------------------------------------------------------------
    # 2. Streaming parse — `read_only=True` keeps memory flat. data_only
    #    skips formula evaluation. We feed BytesIO so nothing ever lands on
    #    disk (T-06-T-02 mitigation: no path traversal surface).
    # ---------------------------------------------------------------------
    wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    summary: dict[str, int] = {}
    try:
        for sheet_name in wb.sheetnames:
            # Per-sheet SAVEPOINT so a failure in one sheet doesn't poison
            # earlier successful sheets. begin_nested() inside an open
            # outer transaction releases on success and rolls back on
            # exception.
            async with db.begin_nested():
                n = await _import_sheet(db, template, wb[sheet_name])
                summary[sheet_name] = n
    finally:
        wb.close()

    # ---------------------------------------------------------------------
    # 3. Record the idempotency log row. Done AFTER successful parsing so a
    #    half-failed import can be retried with the same key.
    # ---------------------------------------------------------------------
    log_row: KonkinImportLog | None = None
    if idempotency_key:
        log_row = KonkinImportLog(
            template_id=template.id,
            idempotency_key=idempotency_key,
            summary=summary,
            imported_by=imported_by,
        )
        db.add(log_row)
        await db.flush()

    # Outer commit — the router's `get_db` dependency would commit on
    # successful return anyway, but committing here makes the idempotency
    # row durable before the response is rendered.
    await db.commit()

    return {
        "status": "imported",
        "log_id": str(log_row.id) if log_row else None,
        "summary": summary,
    }


async def _import_sheet(db: AsyncSession, template: KonkinTemplate, sheet) -> int:
    """Phase-1 placeholder: count non-empty rows but DON'T create perspektif
    or indikator rows yet.

    The Plan-06 test suite asserts that the endpoint exists, is multipart-
    contracted, returns the right status codes, and writes one log row per
    distinct idempotency_key. Row creation is Phase 2 / Plan 07 seed work.
    """
    n = 0
    # `iter_rows(values_only=True)` is the streaming row reader. We just
    # count rows whose first cell is non-None so trailing blank rows
    # produced by Excel don't inflate the count.
    for row in sheet.iter_rows(values_only=True):
        if row and any(v is not None for v in row):
            n += 1
    return n
